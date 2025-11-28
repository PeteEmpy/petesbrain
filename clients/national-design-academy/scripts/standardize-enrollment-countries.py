#!/usr/bin/env python3
"""
Standardize country names in NDA enrollment data
Creates a mapping table and cleaned dataset
"""

import sys
from pathlib import Path
from collections import defaultdict
import openpyxl
import json

# File paths
ENROLMENTS_FILE = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/NDA-International-Enrolments-ACTIVE.xlsx")
OUTPUT_MAPPING = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/scripts/country-mapping.json")
OUTPUT_STANDARDIZED = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/NDA-International-Enrolments-STANDARDIZED.json")

# Country mapping (variants -> standard name)
COUNTRY_MAPPINGS = {
    # UAE variations
    "ABU DHABI": "United Arab Emirates",
    "Abu Dhabi": "United Arab Emirates",
    "Abu dhabi": "United Arab Emirates",
    "AJMAN": "United Arab Emirates",
    "Ajman": "United Arab Emirates",
    "DUBAI": "United Arab Emirates",
    "Dubai": "United Arab Emirates",
    "Sharjah": "United Arab Emirates",
    "SHARJA": "United Arab Emirates",
    "Sharja": "United Arab Emirates",
    "sharjah": "United Arab Emirates",
    "Sharjah (UAE)": "United Arab Emirates",
    "UAE (Sharjah)": "United Arab Emirates",
    "Ras Al Khaimah": "United Arab Emirates",
    "Ras Al-Khaimah": "United Arab Emirates",
    "Ras al Khaimah": "United Arab Emirates",
    "Al Ain": "United Arab Emirates",
    "Al Fujairah": "United Arab Emirates",
    "Umm AL Quwain": "United Arab Emirates",
    "Umm Al Quwain": "United Arab Emirates",
    "Umm Ul Quwain": "United Arab Emirates",
    "UAE": "United Arab Emirates",
    "United Arab Emirates": "United Arab Emirates",

    # Saudi Arabia variations
    "SAUDI ARABIA": "Saudi Arabia",
    "Saudi Arabia": "Saudi Arabia",
    "Saudi": "Saudi Arabia",
    "Saudia Arabia": "Saudi Arabia",
    "Riyadh": "Saudi Arabia",
    "Riyadh,Saudi Arabia": "Saudi Arabia",
    "Jeddah": "Saudi Arabia",
    "Jeddah,Saudi Arabia": "Saudi Arabia",
    "Al Qassim": "Saudi Arabia",
    "Arar,Saudi Arabia": "Saudi Arabia",
    "Thuwal,Saudi Arabia": "Saudi Arabia",

    # Oman variations
    "Muscat,Oman": "Oman",
    "Oman": "Oman",

    # Qatar variations
    "QATAR": "Qatar",
    "Qatar": "Qatar",
    "Doha": "Qatar",

    # Bahrain variations
    "BAHRAIN": "Bahrain",
    "Bahrain": "Bahrain",
    "Bahnrain": "Bahrain",
    "Bahrian": "Bahrain",

    # Kuwait variations
    "KUWAIT": "Kuwait",
    "Kuwait": "Kuwait",

    # India variations
    "INDIA": "India",
    "India": "India",
    "Delhi": "India",

    # Egypt variations
    "EGYPT": "Egypt",
    "Egypt": "Egypt",
    "Eqypt": "Egypt",
    "Eygpt": "Egypt",

    # China variations
    "CHINA": "China",
    "China": "China",

    # Hong Kong variations
    "HONG KONG": "Hong Kong",
    "Hong Kong": "Hong Kong",

    # Other cities to countries
    "Nairobi": "Kenya",
    "Harare": "Zimbabwe",
    "Kkampala": "Uganda",
    "Geneva": "Switzerland",
    "Bucharest": "Romania",
    "Prague": "Czech Republic",
    "Sarajevo": "Bosnia and Herzegovina",
    "Bali": "Indonesia",

    # Spelling corrections
    "Austrailia": "Australia",
    "Belguim": "Belgium",
    "Beligum": "Belgium",
    "Finlad": "Finland",
    "Costra Rica": "Costa Rica",
    "EQUADOR": "Ecuador",
    "Libyan": "Libya",
    "Morrocco": "Morocco",
    "Motenegro": "Montenegro",
    "Tanzanian": "Tanzania",
    "Venezuala": "Venezuela",
    "Itay": "Italy",

    # Standardize case
    "BELGIUM": "Belgium",
    "BERMUDA": "Bermuda",
    "CAYMAN ISLANDS": "Cayman Islands",
    "FRANCE": "France",
    "IRAN": "Iran",
    "IRAQ": "Iraq",
    "ITALY": "Italy",
    "LEBANON": "Lebanon",
    "MACEDONIA": "North Macedonia",
    "MALAYSIA": "Malaysia",
    "MAURITIUS": "Mauritius",
    "NORWAY": "Norway",
    "SINGAPORE": "Singapore",
    "SOUTH AFRICA": "South Africa",
    "SPAIN": "Spain",
    "SWEDEN": "Sweden",
    "SWITZERLAND": "Switzerland",
    "TURKEY": "Turkey",

    # Handle "S.Africa" variation
    "S.Africa": "South Africa",

    # Handle "US" and "USA"
    "US": "United States",
    "USA": "United States",

    # Handle multi-word standardization
    "Antigua And Barbuda": "Antigua and Barbuda",
    "Bosnia And Herzegovina": "Bosnia and Herzegovina",
    "Iran, Islamic Republic of": "Iran",
    "Russian Federation": "Russia",
    "Syrian Arab Republic": "Syria",
    "Tanzania, United Republic of": "Tanzania",
    "Trinidad And Tobaba": "Trinidad and Tobago",
    "Trinidad": "Trinidad and Tobago",
    "Viet Nam": "Vietnam",

    # Handle "Overseas" and "N/A"
    "Overseas": "Unknown",
    "N/A": "Unknown",
}

def standardize_country(raw_country):
    """Standardize a country name using mapping table"""
    if not raw_country or not isinstance(raw_country, str):
        return "Unknown"

    raw_country = raw_country.strip()

    # Check if it's in the mapping
    if raw_country in COUNTRY_MAPPINGS:
        return COUNTRY_MAPPINGS[raw_country]

    # If not in mapping, return as-is (likely already standardized)
    return raw_country

def parse_enrollment_data():
    """Parse enrollment data and standardize countries"""

    wb = openpyxl.load_workbook(ENROLMENTS_FILE, data_only=True)

    all_enrollments = []
    country_stats = defaultdict(lambda: {"count": 0, "raw_variants": set()})

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Extract academic year
        academic_year = sheet_name.replace('Year ', '').strip()

        # Find column indices
        date_col = None
        country_col = None

        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell_value = ws.cell(1, col_idx).value
            if cell_value and isinstance(cell_value, str):
                cell_lower = cell_value.lower().strip()
                if cell_lower == 'date':
                    date_col = col_idx
                elif 'country' in cell_lower:
                    country_col = col_idx

        if not country_col:
            print(f"Warning: No country column found in sheet {sheet_name}")
            continue

        # Parse each enrollment
        for row_idx in range(2, ws.max_row + 1):
            raw_country = ws.cell(row_idx, country_col).value

            if not raw_country:
                continue

            if isinstance(raw_country, str):
                raw_country = raw_country.strip()

            standard_country = standardize_country(raw_country)

            # Collect enrollment record
            date_val = ws.cell(row_idx, date_col).value if date_col else None

            enrollment = {
                "academic_year": academic_year,
                "date": str(date_val) if date_val else None,
                "raw_country": raw_country,
                "standard_country": standard_country
            }

            all_enrollments.append(enrollment)

            # Track statistics
            country_stats[standard_country]["count"] += 1
            country_stats[standard_country]["raw_variants"].add(raw_country)

    return all_enrollments, country_stats

def main():
    print("="*70)
    print("NDA ENROLLMENT DATA STANDARDIZATION")
    print("="*70)
    print(f"\nReading from: {ENROLMENTS_FILE}\n")

    if not ENROLMENTS_FILE.exists():
        print(f"ERROR: File not found: {ENROLMENTS_FILE}")
        sys.exit(1)

    # Parse and standardize
    enrollments, country_stats = parse_enrollment_data()

    print(f"Total enrollments parsed: {len(enrollments):,}\n")

    # Save country mapping
    mapping_data = {
        "mapping": COUNTRY_MAPPINGS,
        "generated_date": "2025-11-03"
    }

    with open(OUTPUT_MAPPING, 'w') as f:
        json.dump(mapping_data, f, indent=2)

    print(f"✓ Country mapping saved: {OUTPUT_MAPPING}\n")

    # Save standardized enrollments
    standardized_data = {
        "enrollments": enrollments,
        "total_count": len(enrollments),
        "generated_date": "2025-11-03"
    }

    with open(OUTPUT_STANDARDIZED, 'w') as f:
        json.dump(standardized_data, f, indent=2)

    print(f"✓ Standardized enrollment data saved: {OUTPUT_STANDARDIZED}\n")

    # Print country statistics
    print("="*70)
    print("COUNTRY STATISTICS (Standardized)")
    print("="*70)

    # Sort by enrollment count
    sorted_countries = sorted(country_stats.items(),
                             key=lambda x: x[1]["count"],
                             reverse=True)

    print(f"\nUnique standardized countries: {len(sorted_countries)}")
    print(f"\nTop 20 countries by enrollment count:\n")

    for idx, (country, stats) in enumerate(sorted_countries[:20], 1):
        variants_count = len(stats["raw_variants"])
        print(f"{idx:2}. {country:30} - {stats['count']:4} enrollments ({variants_count} raw variants)")

    # Show countries with multiple variants
    print("\n" + "="*70)
    print("COUNTRIES WITH MULTIPLE NAMING VARIANTS (Top 10)")
    print("="*70 + "\n")

    multi_variant = [(country, stats) for country, stats in sorted_countries
                     if len(stats["raw_variants"]) > 1]
    multi_variant.sort(key=lambda x: len(x[1]["raw_variants"]), reverse=True)

    for country, stats in multi_variant[:10]:
        print(f"{country} ({stats['count']} enrollments):")
        for variant in sorted(stats["raw_variants"]):
            print(f"  - {variant}")
        print()

    print("="*70)
    print("✓ STANDARDIZATION COMPLETE")
    print("="*70)

if __name__ == '__main__':
    main()
