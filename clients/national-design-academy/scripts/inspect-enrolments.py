#!/usr/bin/env python3
"""
Inspect the structure of the NDA International Enrolments file
"""

import openpyxl
from pathlib import Path

ENROLMENTS_FILE = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/NDA-International-Enrolments-ACTIVE.xlsx")

wb = openpyxl.load_workbook(ENROLMENTS_FILE, data_only=True)

print(f"Sheets: {wb.sheetnames}\n")

# Look at first sheet in detail
ws = wb[wb.sheetnames[0]]

print(f"=== Sheet: {wb.sheetnames[0]} ===")
print(f"Dimensions: {ws.dimensions}\n")

# Print first 5 rows and up to 15 columns to understand structure
print("First 5 rows:")
for row_idx in range(1, min(6, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(16, ws.max_column + 1)):
        cell = ws.cell(row_idx, col_idx)
        value = str(cell.value)[:30] if cell.value else ""
        row_data.append(value)
    print(f"Row {row_idx}: {' | '.join(row_data)}")

print("\n\nColumn headers (row 1):")
for col_idx in range(1, min(20, ws.max_column + 1)):
    header = ws.cell(1, col_idx).value
    if header:
        print(f"Col {col_idx}: {header}")
