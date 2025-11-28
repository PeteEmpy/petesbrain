#!/usr/bin/env python3
"""
Create professional line chart for NDA International Enrolments
Optimized for browser viewing
"""

import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import rcParams

# File paths
ENROLMENTS_FILE = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/NDA-International-Enrolments-ACTIVE.xlsx")
OUTPUT_CHART = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/nda-international-enrolments-professional.png")

# Professional styling
rcParams['font.family'] = 'sans-serif'
rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'DejaVu Sans']
rcParams['font.size'] = 11
rcParams['axes.linewidth'] = 1.5

def parse_excel_data():
    """Parse Excel file and extract monthly enrolment counts"""

    wb = openpyxl.load_workbook(ENROLMENTS_FILE, data_only=True)
    data_by_year = defaultdict(lambda: defaultdict(int))

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find "Date" column (column 2 based on structure)
        header_row = 1
        date_col = 2

        # Look for exact "date" header to confirm
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(1, col_idx).value
            if cell_value and isinstance(cell_value, str):
                if cell_value.lower().strip() == 'date':
                    date_col = col_idx
                    break

        # Extract academic year from sheet name
        academic_year = sheet_name.replace('Year ', '').strip()

        # Count enrolments by month
        monthly_counts = defaultdict(int)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            date_cell = ws.cell(row_idx, date_col).value

            if not date_cell:
                continue

            try:
                if isinstance(date_cell, datetime):
                    enrol_date = date_cell
                elif isinstance(date_cell, str):
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            enrol_date = datetime.strptime(date_cell.strip(), fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                else:
                    continue

                month_key = enrol_date.strftime('%Y-%m')
                monthly_counts[month_key] += 1

            except Exception:
                continue

        if monthly_counts:
            data_by_year[academic_year] = dict(monthly_counts)

    return data_by_year

def create_professional_chart(data_by_year):
    """Create professional line chart for browser viewing"""

    # Professional color palette (colorblind-friendly)
    colors = [
        '#0173B2',  # Blue
        '#DE8F05',  # Orange
        '#029E73',  # Green
        '#CC78BC',  # Purple
        '#CA9161',  # Brown
        '#949494',  # Grey
        '#ECE133',  # Yellow
        '#56B4E9',  # Light Blue
    ]

    # Create figure with high DPI
    fig, ax = plt.subplots(figsize=(16, 9), dpi=120)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('#FAFAFA')

    # Sort years and plot each one
    sorted_years = sorted(data_by_year.keys(), reverse=False)

    for idx, year in enumerate(sorted_years):
        monthly_data = data_by_year[year]

        if not monthly_data:
            continue

        # Sort months chronologically
        sorted_months = sorted(monthly_data.keys())
        dates = [datetime.strptime(m, '%Y-%m') for m in sorted_months]
        counts = [monthly_data[m] for m in sorted_months]

        color = colors[idx % len(colors)]

        # Plot line with markers
        ax.plot(dates, counts,
                marker='o',
                linewidth=2.5,
                markersize=7,
                label=year,
                color=color,
                markeredgewidth=1.5,
                markeredgecolor='white',
                alpha=0.9)

    # Styling
    ax.set_xlabel('Month', fontsize=14, fontweight='600', labelpad=10)
    ax.set_ylabel('Number of Enrolments', fontsize=14, fontweight='600', labelpad=10)
    ax.set_title('NDA International Enrolments by Month and Academic Year',
                 fontsize=18, fontweight='700', pad=20)

    # Format x-axis
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=3))
    plt.xticks(rotation=45, ha='right')

    # Grid styling
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.8, color='#CCCCCC')
    ax.set_axisbelow(True)

    # Legend styling
    legend = ax.legend(loc='upper left',
                      framealpha=0.95,
                      fontsize=11,
                      title='Academic Year',
                      title_fontsize=12,
                      edgecolor='#CCCCCC',
                      fancybox=True,
                      shadow=False)
    legend.get_title().set_fontweight('600')

    # Spine styling
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    for spine in ['bottom', 'left']:
        ax.spines[spine].set_color('#CCCCCC')
        ax.spines[spine].set_linewidth(1.5)

    # Tick styling
    ax.tick_params(axis='both', which='major', labelsize=10,
                   colors='#333333', length=6, width=1.5)

    # Add subtle branding
    fig.text(0.99, 0.01, 'National Design Academy | Enrolment Analysis',
             ha='right', va='bottom', fontsize=9, color='#999999',
             style='italic')

    # Tight layout with padding
    plt.tight_layout(pad=1.5)

    # Save with high quality
    plt.savefig(OUTPUT_CHART, dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')

    print(f"\nProfessional chart saved to: {OUTPUT_CHART}")
    print(f"Dimensions: 2400x1350 pixels (16:9 ratio)")
    print(f"Optimized for browser viewing at 150 DPI")

    return OUTPUT_CHART

def print_summary(data_by_year):
    """Print summary with current year highlighted"""
    print("\n" + "="*70)
    print("ENROLMENT SUMMARY BY ACADEMIC YEAR")
    print("="*70)

    for year in sorted(data_by_year.keys(), reverse=False):
        monthly_data = data_by_year[year]
        total = sum(monthly_data.values())
        months = len(monthly_data)
        avg = total / months if months > 0 else 0

        marker = " ← CURRENT" if year == "2025-26" else ""

        print(f"\n{year}{marker}")
        print(f"  Total Enrolments: {total:,}")
        print(f"  Months Active: {months}")
        print(f"  Average/Month: {avg:.1f}")

        if monthly_data:
            best_month = max(monthly_data.items(), key=lambda x: x[1])
            print(f"  Peak Month: {datetime.strptime(best_month[0], '%Y-%m').strftime('%b %Y')} ({best_month[1]} enrolments)")

if __name__ == '__main__':
    print("Creating professional chart for NDA International Enrolments...")
    print(f"Reading from: {ENROLMENTS_FILE}\n")

    if not ENROLMENTS_FILE.exists():
        print(f"ERROR: File not found: {ENROLMENTS_FILE}")
        sys.exit(1)

    # Parse data
    data = parse_excel_data()

    if not data:
        print("ERROR: No data found in Excel file")
        sys.exit(1)

    # Print summary
    print_summary(data)

    # Create professional chart
    chart_path = create_professional_chart(data)

    print("\n✓ Chart created successfully!")
    print(f"  View at: {chart_path}")
