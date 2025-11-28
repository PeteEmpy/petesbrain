import openpyxl
from datetime import datetime
from collections import defaultdict

# Load both spreadsheets
uk_wb = openpyxl.load_workbook('enrolments/NDA-UK-Enrolments-ACTIVE.xlsx')
intl_wb = openpyxl.load_workbook('enrolments/NDA-International-Enrolments-ACTIVE.xlsx')

uk_ws = uk_wb.active
intl_ws = intl_wb.active

# Get headers
uk_headers = [cell.value for cell in uk_ws[1]]
intl_headers = [cell.value for cell in intl_ws[1]]

print("UK Headers:", uk_headers)
print("\nInternational Headers:", intl_headers)

# Find the date and revenue columns
uk_date_col = uk_headers.index('Enrolled') + 1 if 'Enrolled' in uk_headers else None
uk_revenue_col = uk_headers.index('Order Value') + 1 if 'Order Value' in uk_headers else None

intl_date_col = intl_headers.index('Enrolled') + 1 if 'Enrolled' in intl_headers else None
intl_revenue_col = intl_headers.index('Order Value') + 1 if 'Order Value' in intl_headers else None

print(f"\nUK: Date column {uk_date_col}, Revenue column {uk_revenue_col}")
print(f"International: Date column {intl_date_col}, Revenue column {intl_revenue_col}")

# Collect all data
uk_data = []
for row in uk_ws.iter_rows(min_row=2, values_only=True):
    if row[uk_date_col-1] and row[uk_revenue_col-1]:
        date = row[uk_date_col-1]
        if isinstance(date, datetime):
            date_str = date.strftime('%Y-%m-%d')
        else:
            date_str = str(date)
        
        revenue = row[uk_revenue_col-1]
        try:
            revenue_float = float(str(revenue).replace('£', '').replace(',', ''))
        except:
            revenue_float = 0.0
        
        uk_data.append({
            'date': date_str,
            'revenue': revenue_float,
            'market': 'UK'
        })

intl_data = []
for row in intl_ws.iter_rows(min_row=2, values_only=True):
    if row[intl_date_col-1] and row[intl_revenue_col-1]:
        date = row[intl_date_col-1]
        if isinstance(date, datetime):
            date_str = date.strftime('%Y-%m-%d')
        else:
            date_str = str(date)
        
        revenue = row[intl_revenue_col-1]
        try:
            revenue_float = float(str(revenue).replace('£', '').replace(',', ''))
        except:
            revenue_float = 0.0
        
        intl_data.append({
            'date': date_str,
            'revenue': revenue_float,
            'market': 'International'
        })

print(f"\n\nTotal UK enrollments: {len(uk_data)}")
print(f"Total International enrollments: {len(intl_data)}")
print(f"Total ALL enrollments: {len(uk_data) + len(intl_data)}")

# Calculate totals
uk_total_revenue = sum(d['revenue'] for d in uk_data)
intl_total_revenue = sum(d['revenue'] for d in intl_data)
total_revenue = uk_total_revenue + intl_total_revenue

uk_avg_revenue = uk_total_revenue / len(uk_data) if uk_data else 0
intl_avg_revenue = intl_total_revenue / len(intl_data) if intl_data else 0
overall_avg_revenue = total_revenue / (len(uk_data) + len(intl_data)) if (uk_data or intl_data) else 0

print("\n" + "="*70)
print("REVENUE DATA FROM ENROLLMENT SPREADSHEETS")
print("="*70)

print(f"\nUK STUDENTS:")
print(f"  Enrollments: {len(uk_data)}")
print(f"  Total Revenue: £{uk_total_revenue:,.2f}")
print(f"  Average Revenue per Enrollment: £{uk_avg_revenue:,.2f}")

print(f"\nINTERNATIONAL STUDENTS:")
print(f"  Enrollments: {len(intl_data)}")
print(f"  Total Revenue: £{intl_total_revenue:,.2f}")
print(f"  Average Revenue per Enrollment: £{intl_avg_revenue:,.2f}")

print(f"\nOVERALL:")
print(f"  Total Enrollments: {len(uk_data) + len(intl_data)}")
print(f"  Total Revenue: £{total_revenue:,.2f}")
print(f"  Average Revenue per Enrollment: £{overall_avg_revenue:,.2f}")

print("\n" + "="*70)

# October-specific data
print("\nOCTOBER 2025 BREAKDOWN:")
print("="*70)

uk_oct = [d for d in uk_data if d['date'].startswith('2025-10')]
intl_oct = [d for d in intl_data if d['date'].startswith('2025-10')]

uk_oct_revenue = sum(d['revenue'] for d in uk_oct)
intl_oct_revenue = sum(d['revenue'] for d in intl_oct)
oct_total_revenue = uk_oct_revenue + intl_oct_revenue

uk_oct_avg = uk_oct_revenue / len(uk_oct) if uk_oct else 0
intl_oct_avg = intl_oct_revenue / len(intl_oct) if intl_oct else 0
oct_avg = oct_total_revenue / (len(uk_oct) + len(intl_oct)) if (uk_oct or intl_oct) else 0

print(f"\nUK STUDENTS (October):")
print(f"  Enrollments: {len(uk_oct)}")
print(f"  Total Revenue: £{uk_oct_revenue:,.2f}")
print(f"  Average Revenue per Enrollment: £{uk_oct_avg:,.2f}")

print(f"\nINTERNATIONAL STUDENTS (October):")
print(f"  Enrollments: {len(intl_oct)}")
print(f"  Total Revenue: £{intl_oct_revenue:,.2f}")
print(f"  Average Revenue per Enrollment: £{intl_oct_avg:,.2f}")

print(f"\nOVERALL (October):")
print(f"  Total Enrollments: {len(uk_oct) + len(intl_oct)}")
print(f"  Total Revenue: £{oct_total_revenue:,.2f}")
print(f"  Average Revenue per Enrollment: £{oct_avg:,.2f}")

print("\n" + "="*70)

# Monthly breakdown
print("\nMONTHLY BREAKDOWN (All Students):")
print("="*70)

monthly_data = defaultdict(lambda: {'uk_count': 0, 'intl_count': 0, 'uk_revenue': 0, 'intl_revenue': 0})

for d in uk_data:
    month = d['date'][:7]  # YYYY-MM
    monthly_data[month]['uk_count'] += 1
    monthly_data[month]['uk_revenue'] += d['revenue']

for d in intl_data:
    month = d['date'][:7]  # YYYY-MM
    monthly_data[month]['intl_count'] += 1
    monthly_data[month]['intl_revenue'] += d['revenue']

print(f"\n{'Month':<12} {'UK':>8} {'Intl':>8} {'Total':>8} {'UK Rev':>12} {'Intl Rev':>12} {'Total Rev':>12}")
print("-" * 85)

for month in sorted(monthly_data.keys()):
    data = monthly_data[month]
    total_count = data['uk_count'] + data['intl_count']
    total_revenue = data['uk_revenue'] + data['intl_revenue']
    
    print(f"{month:<12} {data['uk_count']:>8} {data['intl_count']:>8} {total_count:>8} "
          f"£{data['uk_revenue']:>10,.0f} £{data['intl_revenue']:>10,.0f} £{total_revenue:>10,.0f}")

print("="*85)

