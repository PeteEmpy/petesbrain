#!/usr/bin/env python3
"""
Create line chart comparing academic years by calendar month
X-axis: 12 months (Jan-Dec)
Each line: One academic year
"""

import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib import rcParams
from PIL import Image

# File paths
ENROLMENTS_FILE = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/NDA-International-Enrolments-ACTIVE.xlsx")
OUTPUT_CHART = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/nda-international-by-month-comparison.png")

# Professional styling
rcParams['font.family'] = 'sans-serif'
rcParams['font.sans-serif'] = ['Arial', 'Helvetica', 'DejaVu Sans']
rcParams['font.size'] = 11
rcParams['axes.linewidth'] = 1.5

def parse_excel_data():
    """Parse Excel and group by academic year and calendar month"""

    wb = openpyxl.load_workbook(ENROLMENTS_FILE, data_only=True)
    data_by_year = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Find "Date" column
        date_col = 2
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(1, col_idx).value
            if cell_value and isinstance(cell_value, str):
                if cell_value.lower().strip() == 'date':
                    date_col = col_idx
                    break

        # Extract academic year
        academic_year = sheet_name.replace('Year ', '').strip()

        # Count by calendar month (1-12)
        monthly_counts = defaultdict(int)

        for row_idx in range(2, ws.max_row + 1):
            date_cell = ws.cell(row_idx, date_col).value

            if not date_cell:
                continue

            try:
                if isinstance(date_cell, datetime):
                    enrol_date = date_cell
                elif isinstance(date_cell, str):
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                        try:
                            enrol_date = datetime.strptime(date_cell.strip(), fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                else:
                    continue

                # Use month number (1-12)
                month_num = enrol_date.month
                monthly_counts[month_num] += 1

            except Exception:
                continue

        if monthly_counts:
            data_by_year[academic_year] = dict(monthly_counts)

    return data_by_year

def create_comparison_chart(data_by_year):
    """Create chart with months on X-axis and one line per year"""

    # Bold, distinct colors for 4 years
    colors = [
        '#0066CC',  # Strong Blue
        '#FF6B35',  # Vibrant Orange
        '#2ECC71',  # Bright Green
        '#9B59B6',  # Rich Purple
    ]

    # Create larger figure with more white space
    fig, ax = plt.subplots(figsize=(18, 10), dpi=120)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')

    # Month labels
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    month_numbers = list(range(1, 13))

    # Sort years and take only last 4
    sorted_years = sorted(data_by_year.keys(), reverse=False)[-4:]

    # Plot each year (oldest to newest, so newest is on top)
    for idx, year in enumerate(sorted_years):
        monthly_data = data_by_year[year]

        # Create array of 12 months, with 0 for months with no data
        counts = [monthly_data.get(month, 0) for month in month_numbers]

        color = colors[idx % len(colors)]

        # Plot line with thicker, more visible styling
        # Use idx for zorder so newer years are on top (higher zorder)
        ax.plot(month_numbers, counts,
                marker='o',
                linewidth=3.5,
                markersize=10,  # Slightly larger markers
                label=year,
                color=color,
                markeredgewidth=2,
                markeredgecolor='white',
                alpha=0.95,
                zorder=idx + 1)  # Newer years get higher zorder (on top)

    # Professional styling with larger fonts and better spacing
    ax.set_xlabel('Month', fontsize=17, fontweight='600', labelpad=15, color='#333333')
    ax.set_ylabel('Number of Enrolments', fontsize=17, fontweight='600', labelpad=15, color='#333333')

    # Main title with subtitle
    ax.set_title('National Design Academy\nInternational Enrolments by Month',
                 fontsize=22, fontweight='700', pad=30, color='#333333', loc='center')

    # Set X-axis to show month names with better spacing
    ax.set_xticks(month_numbers)
    ax.set_xticklabels(month_names, fontsize=13)

    # Lighter, more subtle grid
    ax.grid(True, alpha=0.2, linestyle='--', linewidth=0.8, color='#DDDDDD')
    ax.set_axisbelow(True)

    # Legend styling - larger and clearer
    legend = ax.legend(loc='upper right',
                      framealpha=0.98,
                      fontsize=14,
                      title='Academic Year',
                      title_fontsize=15,
                      edgecolor='#AAAAAA',
                      fancybox=True,
                      shadow=True,
                      borderpad=1.2,
                      labelspacing=1.0)
    legend.get_title().set_fontweight('700')

    # Clean spine styling
    for spine in ['top', 'right']:
        ax.spines[spine].set_visible(False)
    for spine in ['bottom', 'left']:
        ax.spines[spine].set_color('#999999')
        ax.spines[spine].set_linewidth(2)

    # Larger tick labels
    ax.tick_params(axis='both', which='major', labelsize=12,
                   colors='#333333', length=8, width=2)

    # Set Y-axis to start at 0
    ax.set_ylim(bottom=0)

    # Add Roksys logo (bottom-right, subtle)
    logo_path = Path('/Users/administrator/Documents/PetesBrain/shared/assets/branding/roksys-logo-200x50.png')
    if logo_path.exists():
        try:
            logo = Image.open(logo_path)
            # Scale logo smaller (40% of original)
            new_width = int(logo.width * 0.4)
            new_height = int(logo.height * 0.4)
            logo_resized = logo.resize((new_width, new_height), Image.Resampling.LANCZOS)

            # Position in bottom-right corner
            fig.figimage(logo_resized,
                        xo=fig.bbox.xmax - new_width - 15,
                        yo=15,
                        alpha=0.7,
                        zorder=1)
        except Exception as e:
            print(f"Note: Could not add logo: {e}")

    # Tight layout with extra padding at bottom for branding
    plt.tight_layout(pad=2.0, rect=[0, 0.02, 1, 1])

    # Save
    plt.savefig(OUTPUT_CHART, dpi=150, bbox_inches='tight',
                facecolor='white', edgecolor='none')

    print(f"\nChart saved to: {OUTPUT_CHART}")
    print(f"Format: 12 months on X-axis, one line per academic year")

    return OUTPUT_CHART

def print_summary(data_by_year):
    """Print summary by year"""
    print("\n" + "="*70)
    print("ENROLMENT SUMMARY BY ACADEMIC YEAR")
    print("="*70)

    for year in sorted(data_by_year.keys(), reverse=False):
        monthly_data = data_by_year[year]
        total = sum(monthly_data.values())

        marker = " ← CURRENT" if year == "2025-26" else ""

        print(f"\n{year}{marker}")
        print(f"  Total: {total:,}")

        # Show month-by-month
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        counts_str = []
        for i in range(1, 13):
            count = monthly_data.get(i, 0)
            if count > 0:
                counts_str.append(f"{month_names[i-1]}:{count}")

        print(f"  By month: {', '.join(counts_str)}")

if __name__ == '__main__':
    print("Creating monthly comparison chart...")
    print(f"Reading from: {ENROLMENTS_FILE}\n")

    if not ENROLMENTS_FILE.exists():
        print(f"ERROR: File not found: {ENROLMENTS_FILE}")
        sys.exit(1)

    # Parse data
    data = parse_excel_data()

    if not data:
        print("ERROR: No data found")
        sys.exit(1)

    # Print summary
    print_summary(data)

    # Create chart
    chart_path = create_comparison_chart(data)

    print("\n✓ Chart created successfully!")
