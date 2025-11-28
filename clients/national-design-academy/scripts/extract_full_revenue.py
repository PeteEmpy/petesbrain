import openpyxl
from datetime import datetime

# Load both spreadsheets
uk_wb = openpyxl.load_workbook('enrolments/NDA-UK-Enrolments-ACTIVE.xlsx')
intl_wb = openpyxl.load_workbook('enrolments/NDA-International-Enrolments-ACTIVE.xlsx')

uk_ws = uk_wb.active
intl_ws = intl_wb.active

# Get headers
uk_headers = [cell.value for cell in uk_ws[1]]
intl_headers = [cell.value for cell in intl_ws[1]]

# Column indices
uk_date_col = 2  # Date
uk_fee_col = 8   # Full Course fee
intl_date_col = 2
intl_fee_col = 8

# USD to GBP rate
usd_to_gbp = 0.79

# Extract UK data
uk_data = []
for row in uk_ws.iter_rows(min_row=2, values_only=True):
    if row[uk_date_col-1] and row[uk_fee_col-1]:
        date = row[uk_date_col-1]
        if isinstance(date, datetime):
            date_str = date.strftime('%Y-%m-%d')
        else:
            date_str = str(date)
        
        fee = row[uk_fee_col-1]
        try:
            # UK fees are in GBP
            fee_gbp = float(str(fee).replace('£', '').replace(',', '').strip())
        except:
            fee_gbp = 0.0
        
        uk_data.append({'date': date_str, 'revenue': fee_gbp})

# Extract International data
intl_data = []
for row in intl_ws.iter_rows(min_row=2, values_only=True):
    if row[intl_date_col-1] and row[intl_fee_col-1]:
        date = row[intl_date_col-1]
        if isinstance(date, datetime):
            date_str = date.strftime('%Y-%m-%d')
        else:
            date_str = str(date)
        
        fee = row[intl_fee_col-1]
        try:
            # International fees are in USD
            fee_usd = float(str(fee).replace('$', '').replace(',', '').strip())
            fee_gbp = fee_usd * usd_to_gbp
        except:
            fee_gbp = 0.0
        
        intl_data.append({'date': date_str, 'revenue': fee_gbp})

# Calculate YTD totals
uk_total = sum(d['revenue'] for d in uk_data)
intl_total = sum(d['revenue'] for d in intl_data)
overall_total = uk_total + intl_total

uk_count = len(uk_data)
intl_count = len(intl_data)
total_count = uk_count + intl_count

print("="*70)
print("YTD 2025-26 REVENUE (Aug - Oct 2025) - FULL COURSE FEES")
print("="*70)

print(f"\nUK STUDENTS:")
print(f"  Enrollments: {uk_count}")
print(f"  Total Revenue: £{uk_total:,.2f}")
print(f"  Revenue per Enrollment: £{uk_total/uk_count:,.2f}")

print(f"\nINTERNATIONAL STUDENTS:")
print(f"  Enrollments: {intl_count}")
print(f"  Total Revenue: £{intl_total:,.2f}")
print(f"  Revenue per Enrollment: £{intl_total/intl_count:,.2f}")

print(f"\nOVERALL:")
print(f"  Total Enrollments: {total_count}")
print(f"  Total Revenue: £{overall_total:,.2f}")
print(f"  Revenue per Enrollment: £{overall_total/total_count:,.2f}")

print("\n" + "="*70)

# October breakdown
uk_oct = [d for d in uk_data if d['date'].startswith('2025-10')]
intl_oct = [d for d in intl_data if d['date'].startswith('2025-10')]

uk_oct_total = sum(d['revenue'] for d in uk_oct)
intl_oct_total = sum(d['revenue'] for d in intl_oct)
oct_total = uk_oct_total + intl_oct_total

print("\nOCTOBER 2025 ONLY:")
print("="*70)

print(f"\nUK STUDENTS:")
print(f"  Enrollments: {len(uk_oct)}")
print(f"  Total Revenue: £{uk_oct_total:,.2f}")
print(f"  Revenue per Enrollment: £{uk_oct_total/len(uk_oct):,.2f}")

print(f"\nINTERNATIONAL STUDENTS:")
print(f"  Enrollments: {len(intl_oct)}")
print(f"  Total Revenue: £{intl_oct_total:,.2f}")
print(f"  Revenue per Enrollment: £{intl_oct_total/len(intl_oct):,.2f}")

print(f"\nOVERALL:")
print(f"  Total Enrollments: {len(uk_oct) + len(intl_oct)}")
print(f"  Total Revenue: £{oct_total:,.2f}")
print(f"  Revenue per Enrollment: £{oct_total/(len(uk_oct) + len(intl_oct)):,.2f}")

print("\n" + "="*70)

# Compare to Paul's financial data
print("\nCOMPARISON TO PAUL'S FINANCIAL DATA (YTD 2025-26):")
print("="*70)

paul_uk_income = 390083
paul_dubai_income = 404160
paul_total = 794243

print(f"\n{'Source':<30} {'UK':<15} {'Intl':<15} {'Total':<15}")
print("-" * 70)
print(f"{'Spreadsheet (full fees):':<30} £{uk_total:>12,.0f} £{intl_total:>12,.0f} £{overall_total:>12,.0f}")
print(f"{'Paul (financial data):':<30} £{paul_uk_income:>12,} £{paul_dubai_income:>12,} £{paul_total:>12,}")
print(f"{'Difference:':<30} £{uk_total-paul_uk_income:>12,.0f} £{intl_total-paul_dubai_income:>12,.0f} £{overall_total-paul_total:>12,.0f}")

print("\n" + "="*70)

# October with Google Ads spend
oct_intl_spend = 22254.96

print("\nOCTOBER 2025 INTERNATIONAL WITH GOOGLE ADS SPEND:")
print("="*70)
print(f"  Enrollments: {len(intl_oct)}")
print(f"  Revenue: £{intl_oct_total:,.2f}")
print(f"  Google Ads Spend: £{oct_intl_spend:,.2f}")
print(f"  Cost per Enrollment: £{oct_intl_spend/len(intl_oct):,.2f}")
print(f"  Revenue per Enrollment: £{intl_oct_total/len(intl_oct):,.2f}")
print(f"  ROAS: {(intl_oct_total/oct_intl_spend)*100:.0f}%")
print("="*70)

