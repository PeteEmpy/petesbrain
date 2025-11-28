#!/usr/bin/env python3
"""
Analyze NDA International Enrolments and create visualization
"""

import sys
import json
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

# File paths
ENROLMENTS_FILE = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/NDA-International-Enrolments-ACTIVE.xlsx")
OUTPUT_CHART = Path("/Users/administrator/Documents/PetesBrain/clients/national-design-academy/enrolments/international-enrolments-by-month.png")

def parse_excel_data():
    """Parse the Excel file and extract enrolment data by month and academic year"""

    wb = openpyxl.load_workbook(ENROLMENTS_FILE, data_only=True)

    # Data structure: {academic_year: {month: count}}
    data_by_year = defaultdict(lambda: defaultdict(int))

    print(f"Available sheets: {wb.sheetnames}")

    # Process each sheet (each sheet should be an academic year)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        print(f"\n=== Processing sheet: {sheet_name} ===")
        print(f"Dimensions: {ws.dimensions}")

        # Find the header row
        header_row = None
        date_col = None

        # Look for "Date" column specifically (not "W/C Date")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell_value = ws.cell(row_idx, col_idx).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower().strip()
                    # Look for exact "date" column, not "w/c date"
                    if cell_lower == 'date':
                        header_row = row_idx
                        date_col = col_idx
                        print(f"Found date column '{cell_value}' at row {row_idx}, col {col_idx}")
                        break
            if header_row:
                break

        if not header_row or not date_col:
            print(f"Warning: Could not find 'Date' column in sheet {sheet_name}, trying column 2")
            # Based on inspection, Date is typically column 2
            header_row = 1
            date_col = 2

        # Extract academic year from sheet name
        # Common patterns: "2024-25", "24-25", "AY 2024-25", etc.
        academic_year = sheet_name.strip()

        # Count enrolments by month
        monthly_counts = defaultdict(int)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            date_cell = ws.cell(row_idx, date_col).value

            if not date_cell:
                continue

            # Handle different date formats
            try:
                if isinstance(date_cell, datetime):
                    enrol_date = date_cell
                elif isinstance(date_cell, str):
                    # Try parsing common date formats
                    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                        try:
                            enrol_date = datetime.strptime(date_cell.strip(), fmt)
                            break
                        except ValueError:
                            continue
                    else:
                        continue
                else:
                    continue

                # Create month key (YYYY-MM)
                month_key = enrol_date.strftime('%Y-%m')
                monthly_counts[month_key] += 1

            except Exception as e:
                continue

        print(f"Found {sum(monthly_counts.values())} enrolments across {len(monthly_counts)} months")

        # Store in main data structure
        if monthly_counts:
            data_by_year[academic_year] = dict(monthly_counts)

    return data_by_year

def create_chart(data_by_year):
    """Create line chart showing enrolments by month for each academic year"""

    if not data_by_year:
        print("No data to plot!")
        return

    # Set up the plot
    fig, ax = plt.subplots(figsize=(14, 8))

    # Color palette for different years
    colors = ['#2E86AB', '#A23B72', '#F18F01', '#C73E1D', '#6A994E', '#BC4B51']

    # Plot each academic year
    for idx, (year, monthly_data) in enumerate(sorted(data_by_year.items())):
        if not monthly_data:
            continue

        # Sort months chronologically
        sorted_months = sorted(monthly_data.keys())
        dates = [datetime.strptime(m, '%Y-%m') for m in sorted_months]
        counts = [monthly_data[m] for m in sorted_months]

        color = colors[idx % len(colors)]
        ax.plot(dates, counts, marker='o', linewidth=2, markersize=6,
                label=f"Academic Year {year}", color=color)

    # Formatting
    ax.set_xlabel('Month', fontsize=12, fontweight='bold')
    ax.set_ylabel('Number of Enrolments', fontsize=12, fontweight='bold')
    ax.set_title('NDA International Enrolments by Month and Academic Year',
                 fontsize=14, fontweight='bold', pad=20)

    # Format x-axis to show months nicely
    ax.xaxis.set_major_formatter(mdates.DateFormatter('%b %Y'))
    ax.xaxis.set_major_locator(mdates.MonthLocator(interval=2))
    plt.xticks(rotation=45, ha='right')

    # Grid for easier reading
    ax.grid(True, alpha=0.3, linestyle='--')

    # Legend
    ax.legend(loc='upper left', framealpha=0.9, fontsize=10)

    # Tight layout
    plt.tight_layout()

    # Save
    plt.savefig(OUTPUT_CHART, dpi=300, bbox_inches='tight')
    print(f"\nChart saved to: {OUTPUT_CHART}")

    return OUTPUT_CHART

def print_summary(data_by_year):
    """Print summary statistics"""
    print("\n" + "="*60)
    print("SUMMARY STATISTICS")
    print("="*60)

    for year, monthly_data in sorted(data_by_year.items()):
        total = sum(monthly_data.values())
        months = len(monthly_data)
        avg = total / months if months > 0 else 0

        print(f"\nAcademic Year: {year}")
        print(f"  Total Enrolments: {total}")
        print(f"  Months with Data: {months}")
        print(f"  Average per Month: {avg:.1f}")

        if monthly_data:
            sorted_months = sorted(monthly_data.items(), key=lambda x: x[1], reverse=True)
            best_month = sorted_months[0]
            print(f"  Best Month: {best_month[0]} ({best_month[1]} enrolments)")

if __name__ == '__main__':
    print("Analyzing NDA International Enrolments...")
    print(f"Reading from: {ENROLMENTS_FILE}")

    if not ENROLMENTS_FILE.exists():
        print(f"ERROR: File not found: {ENROLMENTS_FILE}")
        sys.exit(1)

    # Parse data
    data = parse_excel_data()

    if not data:
        print("ERROR: No data found in Excel file")
        sys.exit(1)

    # Print summary
    print_summary(data)

    # Create chart
    chart_path = create_chart(data)

    print("\nDone!")
