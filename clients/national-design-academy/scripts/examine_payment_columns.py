import openpyxl

# Load international spreadsheet
intl_wb = openpyxl.load_workbook('enrolments/NDA-International-Enrolments-ACTIVE.xlsx')
intl_ws = intl_wb.active

# Get headers
headers = [cell.value for cell in intl_ws[1]]

print("Column headers:")
for i, h in enumerate(headers, 1):
    print(f"  {i}. {h}")

print("\n" + "="*80)
print("SAMPLE DATA (First 10 international enrollments):")
print("="*80)

# Show first 10 rows with all payment-related columns
print(f"\n{'Date':<12} {'Country':<15} {'Full Fee':<12} {'Paid Full':<12} {'Initial':<12} {'BACS':<8} {'Checkout':<10}")
print("-" * 85)

for i, row in enumerate(intl_ws.iter_rows(min_row=2, max_row=11, values_only=True), 1):
    date = row[1] if len(row) > 1 else ''
    country = row[6] if len(row) > 6 else ''
    full_fee = row[7] if len(row) > 7 else ''
    paid_full = row[8] if len(row) > 8 else ''
    initial = row[9] if len(row) > 9 else ''
    bacs = row[10] if len(row) > 10 else ''
    checkout = row[11] if len(row) > 11 else ''
    
    print(f"{str(date)[:10]:<12} {str(country):<15} {str(full_fee):<12} {str(paid_full):<12} {str(initial):<12} {str(bacs):<8} {str(checkout):<10}")

print("\n" + "="*80)
print("OCTOBER 2025 INTERNATIONAL STUDENTS:")
print("="*80)

print(f"\n{'Date':<12} {'Country':<15} {'Full Fee':<12} {'Paid Full':<12} {'Initial':<12} {'BACS':<8} {'Checkout':<10}")
print("-" * 85)

oct_count = 0
for row in intl_ws.iter_rows(min_row=2, values_only=True):
    date = row[1] if len(row) > 1 else ''
    if date and '2025-10' in str(date):
        oct_count += 1
        country = row[6] if len(row) > 6 else ''
        full_fee = row[7] if len(row) > 7 else ''
        paid_full = row[8] if len(row) > 8 else ''
        initial = row[9] if len(row) > 9 else ''
        bacs = row[10] if len(row) > 10 else ''
        checkout = row[11] if len(row) > 11 else ''
        
        print(f"{str(date)[:10]:<12} {str(country):<15} {str(full_fee):<12} {str(paid_full):<12} {str(initial):<12} {str(bacs):<8} {str(checkout):<10}")

print(f"\nTotal October international enrollments: {oct_count}")

